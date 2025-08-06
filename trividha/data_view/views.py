from django.shortcuts import render

# Create your views here.
from django.http import HttpResponse
from openpyxl import Workbook
from form.models import participant_details,events,school
from openpyxl.styles import Font, Alignment
from .forms import eventForm
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required



@login_required(login_url='/admin/login/')
def view(request):
    if request.POST:
            wb = Workbook()
            ws = wb.active
            event_name=request.POST.get("event")
            print(event_name)
            ws.title = event_name
            
            ws.merge_cells('A1:C1')  # Merge 3 columns: A to C
            ws['A1'] = event_name
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            part_obj=participant_details.objects.all()
            # Data rows
            ws.append(['PARTICIPANT NAME', 'SCHOOL'])
            ws.append(['', ''])
            for s in part_obj:
                if s.events.name==event_name:



            # Header row
                    
                    print(s)
                    
                    ws.append([
                        s.name,
                        s.school.name
                    ])

                #for adjusting column size
                    for column_cells in ws.columns:
                        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                        col_letter = get_column_letter(column_cells[0].column)
                        ws.column_dimensions[col_letter].width = length + 2

            response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
            response['Content-Disposition'] = f'attachment; filename={event_name}.xlsx'
            wb.save(response)
            return response
    event=events.objects.all()
    return  render(request,"event_data_view.html",{"event":event})


@login_required(login_url='/admin/login/')
def view_school(request):
    if request.POST:
         
            wb = Workbook()
            ws = wb.active
            school_name=request.POST.get("school")
            ws.title = school_name
            
            ws.merge_cells('A1:C1')  # Merge 3 columns: A to C
            ws['A1'] = school_name
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            part_obj=participant_details.objects.all()
            part_obj=participant_details.objects.all()
            ws.append(['PARTICIPANT NAME', 'EVENT'])
            ws.append(['', ''])
        # Data rows
            for s in part_obj:
                if s.school.name==request.POST.get("school"):
                    ws.append([s.name,s.events.name])
                    
                    
            

        #for adjusting column size
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                col_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[col_letter].width = length + 2
            response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            response['Content-Disposition'] = f'attachment; filename={school_name}.xlsx'
            wb.save(response)
            return response
    schools=school.objects.all()
    return  render(request,"data_school.html",{"schools":schools})